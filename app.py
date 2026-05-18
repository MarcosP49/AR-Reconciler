import tkinter as tk
from tkinter import filedialog, messagebox
import os
import sys
import threading

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import transform_ar
from transform_ar import (
    parse_input, write_detail_sheet, write_recap_sheet,
    parse_previous_comments, apply_comments,
    find_ar_sheet, find_recap_sheet, replace_sheet, get_prev_matter_ids,
    AR_SHEET_NAME, RECAP_SHEET_NAME, resolve_blanks,
)
import openpyxl
from openpyxl import Workbook

# ── timekeeper config file (lives next to the exe / script) ───────────────────
if getattr(sys, "frozen", False):
    _BASE = os.path.dirname(sys.executable)
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))
_CONFIG = os.path.join(_BASE, "timekeepers.txt")


def _load_timekeepers():
    """Load names from timekeepers.txt, falling back to the defaults in transform_ar."""
    if os.path.exists(_CONFIG):
        with open(_CONFIG, "r", encoding="utf-8") as f:
            names = [ln.strip() for ln in f if ln.strip()]
        if names:
            return names
    return list(transform_ar.TIMEKEEPER_NAMES)


def _save_timekeepers(names):
    with open(_CONFIG, "w", encoding="utf-8") as f:
        f.write("\n".join(names))

BG       = "#FFFFFF"
BG_LEFT  = "#F7F8F7"
BORDER   = "#E2E2E2"
GREEN    = "#1C6B3A"
GREEN_HV = "#155430"
RED      = "#C0392B"
MUTED    = "#9B9B9B"
DARK     = "#1A1A1A"
XLSX_BG  = "#1C7C44"
F        = "Segoe UI"   

def _count_rows(path):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        n = sum(1 for r in ws.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in r))
        wb.close()
        return n
    except Exception:
        return "?"


def _fmt_money(v):
    return f"${v:,.0f}"


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("AR Reconciler")
        self.geometry("980x600")
        self.resizable(False, False)
        self.configure(bg=BG)

        self._prev_path  = tk.StringVar()
        self._input_path = tk.StringVar()
        self._out_name   = tk.StringVar(value="AR_output.xlsx")
        self._save_dir   = tk.StringVar(value=os.path.expanduser("~"))

        # Load persisted timekeeper names and apply to the engine
        transform_ar.TIMEKEEPER_NAMES = _load_timekeepers()

        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG, height=42)
        top.pack(fill="x")
        top.pack_propagate(False)
        tk.Label(top, text="■  AR Reconciler", font=(F, 12, "bold"),
                 bg=BG, fg=DARK).pack(side="left", padx=18, pady=10)
        self._period_lbl = tk.Label(top, text="", font=(F, 9),
                                    bg=BG, fg=MUTED)
        self._period_lbl.pack(side="right", padx=18)
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        left = tk.Frame(body, bg=BG_LEFT, width=310)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        tk.Frame(body, bg=BORDER, width=1).pack(side="left", fill="y")
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        self._build_left(left)
        self._build_right(right)

    def _build_left(self, p):
        pad = 18

        tk.Label(p, text="INPUTS", font=(F, 8, "bold"),
                 bg=BG_LEFT, fg=MUTED).pack(anchor="w", padx=pad, pady=(18, 6))
        
        self._prev_card  = _FileCard(p, "Previous AR file",  "optional — carries notes forward",
                                     self._prev_path,  self._pick_prev)
        self._prev_card.pack(fill="x", padx=pad, pady=(0, 8))

        self._input_card = _FileCard(p, "Tabs3 input file",  "required — raw export",
                                     self._input_path, self._pick_input)
        self._input_card.pack(fill="x", padx=pad)

        tk.Button(p, text="⚙  Edit timekeeper names",
                  font=(F, 9), bg=BG_LEFT, fg=GREEN, relief="flat",
                  cursor="hand2", anchor="w",
                  command=self._edit_timekeepers
                  ).pack(anchor="w", padx=pad, pady=(10, 0))

        tk.Frame(p, bg=BORDER, height=1).pack(fill="x", padx=pad, pady=(14, 0))
        tk.Label(p, text="OUTPUT", font=(F, 8, "bold"),
                 bg=BG_LEFT, fg=MUTED).pack(anchor="w", padx=pad, pady=(10, 6))

        tk.Label(p, text="Filename", font=(F, 9), bg=BG_LEFT, fg=DARK
                 ).pack(anchor="w", padx=pad)
        tk.Entry(p, textvariable=self._out_name, font=(F, 10), relief="flat",
                 bg=BG, highlightbackground=BORDER, highlightthickness=1
                 ).pack(fill="x", padx=pad, pady=(3, 10))

        tk.Label(p, text="Save to", font=(F, 9), bg=BG_LEFT, fg=DARK
                 ).pack(anchor="w", padx=pad)
        row = tk.Frame(p, bg=BG_LEFT)
        row.pack(fill="x", padx=pad, pady=(3, 0))
        tk.Entry(row, textvariable=self._save_dir, font=(F, 10), relief="flat",
                 bg=BG, highlightbackground=BORDER, highlightthickness=1
                 ).pack(side="left", fill="x", expand=True)
        tk.Button(row, text="…", font=(F, 10), bg=BG, fg=DARK, relief="flat",
                  cursor="hand2", padx=6,
                  command=self._pick_dir).pack(side="left", padx=(4, 0))

        self._status = tk.Label(p, text="", font=(F, 9), bg=BG_LEFT,
                                 fg=MUTED, wraplength=274, justify="left")
        self._status.pack(side="bottom", anchor="w", padx=pad, pady=(0, 4))

        self._gen_btn = tk.Button(p, text="Generate  →",
                                   font=(F, 11, "bold"),
                                   bg=GREEN, fg="white", relief="flat",
                                   cursor="hand2", pady=10,
                                   activebackground=GREEN_HV,
                                   activeforeground="white",
                                   command=self._generate)
        self._gen_btn.pack(side="bottom", fill="x", padx=pad, pady=(0, 14))
        self._gen_btn.bind("<Enter>", lambda _: self._gen_btn.config(bg=GREEN_HV))
        self._gen_btn.bind("<Leave>", lambda _: self._gen_btn.config(bg=GREEN))

    def _edit_timekeepers(self):
        dlg = tk.Toplevel(self)
        dlg.title("Timekeeper Names")
        dlg.geometry("320x420")
        dlg.resizable(False, False)
        dlg.configure(bg=BG)
        dlg.grab_set()  # modal

        tk.Label(dlg, text="Timekeeper Names",
                 font=(F, 12, "bold"), bg=BG, fg=DARK).pack(anchor="w", padx=20, pady=(18, 4))
        tk.Label(dlg, text="One name per line, in the order they\nappear in the input file.",
                 font=(F, 9), bg=BG, fg=MUTED, justify="left").pack(anchor="w", padx=20)

        frame = tk.Frame(dlg, bg=BG, highlightbackground=BORDER, highlightthickness=1)
        frame.pack(fill="both", expand=True, padx=20, pady=12)

        sb = tk.Scrollbar(frame)
        sb.pack(side="right", fill="y")
        txt = tk.Text(frame, font=(F, 10), relief="flat", bg=BG, fg=DARK,
                      yscrollcommand=sb.set, wrap="none", padx=6, pady=6)
        txt.pack(fill="both", expand=True)
        sb.config(command=txt.yview)

        # Populate with current names
        txt.insert("1.0", "\n".join(transform_ar.TIMEKEEPER_NAMES))

        def _save():
            names = [ln.strip() for ln in txt.get("1.0", "end").splitlines() if ln.strip()]
            if not names:
                messagebox.showerror("Empty list", "Please enter at least one name.", parent=dlg)
                return
            transform_ar.TIMEKEEPER_NAMES = names
            _save_timekeepers(names)
            dlg.destroy()

        btn_row = tk.Frame(dlg, bg=BG)
        btn_row.pack(fill="x", padx=20, pady=(0, 16))
        tk.Button(btn_row, text="Cancel", font=(F, 10), bg=BG_LEFT, fg=DARK,
                  relief="flat", cursor="hand2", padx=12, pady=6,
                  command=dlg.destroy).pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text="Save", font=(F, 10, "bold"), bg=GREEN, fg="white",
                  relief="flat", cursor="hand2", padx=12, pady=6,
                  command=_save).pack(side="right")

    def _build_right(self, p):
        hdr = tk.Frame(p, bg=BG)
        hdr.pack(fill="x", padx=24, pady=(16, 0))
        tk.Label(hdr, text="Preview · current AR sheet",
                 font=(F, 13, "bold"), bg=BG, fg=DARK).pack(side="left")
        tk.Label(hdr, text="click Generate to refresh",
                 font=(F, 9), bg=BG, fg=MUTED).pack(side="right", pady=3)

        sf = tk.Frame(p, bg=BG)
        sf.pack(fill="x", padx=24, pady=(14, 0))
        self._s_open    = _Stat(sf, "OPEN",    "—",    DARK)
        self._s_balance = _Stat(sf, "BALANCE", "—",    DARK)
        self._s_new     = _Stat(sf, "NEW",     "—",    GREEN)
        self._s_cleared = _Stat(sf, "CLEARED", "—",    RED)

        tk.Frame(p, bg=BORDER, height=1).pack(fill="x", padx=24, pady=(14, 0))

        COL_W = [14, 26, 12, 9, 9]
        COL_H = ["MATTER", "CLIENT", "BALANCE", "AGE", "STATUS"]
        th = tk.Frame(p, bg=BG)
        th.pack(fill="x", padx=24, pady=(8, 2))
        for h, w in zip(COL_H, COL_W):
            tk.Label(th, text=h, font=(F, 8, "bold"),
                     bg=BG, fg=MUTED, width=w, anchor="w").pack(side="left")

        # scrollable table body
        wrap = tk.Frame(p, bg=BG)
        wrap.pack(fill="both", expand=True, padx=24, pady=(0, 14))
        sb = tk.Scrollbar(wrap, orient="vertical")
        sb.pack(side="right", fill="y")
        self._canvas = tk.Canvas(wrap, bg=BG, highlightthickness=0,
                                  yscrollcommand=sb.set)
        self._canvas.pack(side="left", fill="both", expand=True)
        sb.config(command=self._canvas.yview)
        self._tbody = tk.Frame(self._canvas, bg=BG)
        self._win = self._canvas.create_window((0, 0), window=self._tbody, anchor="nw")
        self._tbody.bind("<Configure>",
            lambda _: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>",
            lambda e: self._canvas.itemconfig(self._win, width=e.width))
        self._canvas.bind_all("<MouseWheel>",
            lambda e: self._canvas.yview_scroll(-1*(e.delta//120), "units"))

        tk.Label(self._tbody, text="No data yet — click Generate",
                 font=(F, 10), bg=BG, fg=MUTED).pack(pady=50)

    def _pick_prev(self):
        p = filedialog.askopenfilename(title="Select previous AR file",
                                       filetypes=[("Excel", "*.xlsx")])
        if p:
            self._prev_path.set(p)
            self._prev_card.set_file(p, _count_rows(p))

    def _pick_input(self):
        p = filedialog.askopenfilename(title="Select Tabs3 input file",
                                       filetypes=[("Excel", "*.xlsx")])
        if p:
            self._input_path.set(p)
            self._input_card.set_file(p, _count_rows(p))
            base = os.path.splitext(os.path.basename(p))[0]
            self._out_name.set(f"{base}_output.xlsx")
            self._save_dir.set(os.path.dirname(p))

    def _pick_dir(self):
        d = filedialog.askdirectory(title="Select output folder")
        if d:
            self._save_dir.set(d)

    #generation
    def _generate(self):
        if not self._input_path.get():
            messagebox.showerror("Missing file", "Please select a Tabs3 input file.")
            return
        self._gen_btn.config(state="disabled", text="Generating…")
        self._status.config(text="Processing…", fg=MUTED)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            inp   = self._input_path.get()
            prev  = self._prev_path.get() or None
            oname = self._out_name.get() or "AR_output.xlsx"
            if not oname.lower().endswith(".xlsx"):
                oname = os.path.splitext(oname)[0] + ".xlsx"
            odir  = self._save_dir.get()
            out   = os.path.join(odir, oname)

            wb_in = openpyxl.load_workbook(inp, data_only=True, read_only=True)
            attorneys, detail_rows, header_info = parse_input(wb_in.active)

            prev_comments = {}
            prev_ids      = set()
            if prev:
                wb_out    = openpyxl.load_workbook(prev)
                prev_ids  = get_prev_matter_ids(wb_out)   # ALL matters, not just commented
                prev_comments = parse_previous_comments(wb_out)
                ar_old   = find_ar_sheet(wb_out)
                rec_old  = find_recap_sheet(wb_out)
                ws_ar    = replace_sheet(wb_out,
                                         ar_old.title  if ar_old  else AR_SHEET_NAME,
                                         ar_old.title  if ar_old  else AR_SHEET_NAME)
                ws_recap = replace_sheet(wb_out,
                                         rec_old.title if rec_old else RECAP_SHEET_NAME,
                                         rec_old.title if rec_old else RECAP_SHEET_NAME)
            else:
                wb_out   = Workbook()
                ws_ar    = wb_out.active
                ws_ar.title = AR_SHEET_NAME
                ws_recap = wb_out.create_sheet(RECAP_SHEET_NAME)

            write_detail_sheet(ws_ar,   attorneys, detail_rows, header_info)
            write_recap_sheet(ws_recap, attorneys, detail_rows, header_info)
            if prev_comments:
                apply_comments(ws_ar, prev_comments)

            wb_out.save(out)

            # Build preview 
            rows, total_bal, new_c = [], 0.0, 0
            new_ids = set()

            for name in attorneys:
                for entry in resolve_blanks(detail_rows.get(name, [])):
                    if entry.get("blank") or entry.get("type") != "matter":
                        continue
                    label  = entry["label"]
                    vals   = entry["values"]
                    bal    = vals.get("Bal Due", 0.0)
                    parts  = label.split(" ", 1)
                    mid    = parts[0]
                    client = parts[1] if len(parts) > 1 else ""
                    age    = next((b for b in reversed(
                                   ["0-30","31-60","61-90","91-120","121-150","151+"])
                                   if vals.get(b, 0) != 0), "")
                    status = "carried" if (prev_ids and mid in prev_ids) else "new"
                    total_bal += bal
                    if status == "new":
                        new_c += 1
                    new_ids.add(mid)
                    rows.append((mid, client, bal, age, status))

            cleared_c = sum(1 for mid in prev_ids if mid not in new_ids)

            self.after(0, lambda o=out, r=rows, tb=total_bal, nc=new_c, cc=cleared_c:
                       self._done(o, r, tb, nc, cc))

        except Exception as ex:
            msg = str(ex)
            self.after(0, lambda m=msg: self._error(m))

    def _done(self, out, rows, total_bal, new_c, cleared_c):
        self._gen_btn.config(state="normal", text="Generate  →")
        self._status.config(text=f"Saved → {os.path.basename(out)}", fg=GREEN)

        self._s_open.update(str(len(rows)))
        self._s_balance.update(_fmt_money(total_bal))
        self._s_new.update(f"+{new_c}")
        self._s_cleared.update(f"-{cleared_c}" if cleared_c else "0")

        for w in self._tbody.winfo_children():
            w.destroy()

        COL_W = [14, 26, 12, 9, 9]
        for i, (mid, client, bal, age, status) in enumerate(rows):
            bg = "#F5F5F5" if i % 2 == 0 else BG
            rf = tk.Frame(self._tbody, bg=bg)
            rf.pack(fill="x")
            sc = GREEN if status == "new" else MUTED
            for txt, w, fg in [
                (mid,              COL_W[0], DARK),
                (client,           COL_W[1], DARK),
                (f"${bal:,.2f}",   COL_W[2], DARK),
                (age,              COL_W[3], MUTED),
                (status,           COL_W[4], sc),
            ]:
                tk.Label(rf, text=txt, font=(F, 9), bg=bg, fg=fg,
                         width=w, anchor="w", pady=5, padx=4).pack(side="left")

    def _error(self, msg):
        self._gen_btn.config(state="normal", text="Generate  →")
        self._status.config(text=f"Error: {msg}", fg=RED)
        messagebox.showerror("Error", msg)


class _FileCard(tk.Frame):
    def __init__(self, parent, label, hint, var, pick_cmd):
        super().__init__(parent, bg="white", relief="flat",
                         highlightbackground=BORDER, highlightthickness=1,
                         cursor="hand2")
        self._var  = var
        self._label = label
        self._hint  = hint

        inner = tk.Frame(self, bg="white")
        inner.pack(fill="both", expand=True, padx=10, pady=8)

        badge = tk.Label(inner, text="XLSX", font=(F, 7, "bold"),
                         bg=XLSX_BG, fg="white", padx=4, pady=2)
        badge.pack(side="left", padx=(0, 10))

        text_col = tk.Frame(inner, bg="white")
        text_col.pack(side="left", fill="both", expand=True)

        self._name = tk.Label(text_col, text=f"No file — {label}",
                               font=(F, 10, "bold"), bg="white", fg=DARK, anchor="w")
        self._name.pack(anchor="w")
        self._hint_lbl = tk.Label(text_col, text=hint, font=(F, 8),
                                   bg="white", fg=MUTED, anchor="w")
        self._hint_lbl.pack(anchor="w")

        x = tk.Button(inner, text="×", font=(F, 12), bg="white", fg=MUTED,
                      relief="flat", cursor="hand2", command=self._clear)
        x.pack(side="right")

        for w in [self, inner, badge, text_col, self._name, self._hint_lbl]:
            w.bind("<Button-1>", lambda _: pick_cmd())

    def set_file(self, path, rows):
        self._var.set(path)
        self._name.config(text=os.path.basename(path))
        self._hint_lbl.config(text=f"{self._label} · {rows} rows")

    def _clear(self):
        self._var.set("")
        self._name.config(text=f"No file — {self._label}")
        self._hint_lbl.config(text=self._hint)


class _Stat(tk.Frame):
    def __init__(self, parent, label, value, fg):
        super().__init__(parent, bg=BG)
        self.pack(side="left", padx=(0, 30))
        tk.Label(self, text=label, font=(F, 8),
                 bg=BG, fg=MUTED).pack(anchor="w")
        self._val = tk.Label(self, text=value, font=(F, 22, "bold"),
                              bg=BG, fg=fg)
        self._val.pack(anchor="w")

    def update(self, text):
        self._val.config(text=text)


if __name__ == "__main__":
    App().mainloop()
