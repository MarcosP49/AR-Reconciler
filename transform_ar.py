import sys
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Patch for openpyxl 3.1.5: some Excel files include an 'id' attr in styles XML.
from openpyxl.styles.cell_style import CellStyle as _CellStyle
_orig_cs_init = _CellStyle.__init__
def _patched_cs_init(self, *args, **kwargs):
    kwargs.pop("id", None)
    _orig_cs_init(self, *args, **kwargs)
_CellStyle.__init__ = _patched_cs_init


#Too many names: extras are ignored. Too few: extra sections use Placeholder
TIMEKEEPER_NAMES = ["Name 1", "Name 2"]

BUCKETS = ["0-30", "31-60", "61-90", "91-120", "121-150", "151+", "Bal Due"]
EXTRA_HEADERS = ["ATTORNEY NOTES", "ACCOUNTING NOTES"]

LABEL_COL = 1        # col A (1-indexed)
BUCKET_START_COL = 2  # col B (1-indexed) — data starts immediately after label

DEBUG = False


def debug(msg):
    if DEBUG:
        print(f"[DEBUG] {msg}", file=sys.stderr)


def get_timekeeper_name(index):
    if index < len(TIMEKEEPER_NAMES):
        return TIMEKEEPER_NAMES[index]
    return f"Placeholder{index + 1}"


def is_blank_row(row):
    return all(cell is None or str(cell).strip() == "" for cell in row)


def detect_column_map(row):
    """Return a full column map if this row contains all 7 bucket headers, else None."""
    mapping = {}
    for col_idx, cell in enumerate(row):
        if cell in BUCKETS:
            mapping[cell] = col_idx
    return mapping if len(mapping) == len(BUCKETS) else None


def strip_m_prefix(label):
    """'2204.12 M Company' → '2204.12 Company'"""
    return re.sub(r'^(\d[\d.]*) M (.+)$', r'\1 \2', label)


def parse_date(text):
    if not text:
        return None
    m = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', str(text))
    if m:
        try:
            return datetime.strptime(m.group(1), "%m/%d/%Y").date()
        except ValueError:
            pass
    return text


def parse_input(ws):
    attorneys = []
    detail_rows = {}
    header_info = {}

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) >= 1:
        r1 = rows[0]
        header_info["date"] = parse_date(r1[0]) if r1[0] else None
        for c in r1[1:]:
            if c and str(c).strip() and "Page:" not in str(c):
                header_info["title"] = str(c).strip()
                break
    if len(rows) >= 2:
        r2 = rows[1]
        for c in r2[1:]:
            if c and str(c).strip():
                header_info["name_label"] = str(c).strip()
                break


    state = "SEEKING_COLUMNS"
    current_name = None
    timekeeper_index = 0
    column_map = {}
    blank_count = 0

    for row in rows:

        # ---- SEEKING_COLUMNS -----------------------------------------------
        if state == "SEEKING_COLUMNS":
            if is_blank_row(row):
                continue
            new_map = detect_column_map(row)
            if new_map:
                column_map = new_map
                current_name = get_timekeeper_name(timekeeper_index)
                attorneys.append(current_name)
                detail_rows[current_name] = []
                debug(f"Columns detected. Starting timekeeper {timekeeper_index}: '{current_name}'")
                debug(f"Column map: {column_map}")
                blank_count = 0
                state = "READING_MATTERS"

        # ---- READING_MATTERS -----------------------------------------------
        elif state == "READING_MATTERS":
            if is_blank_row(row):
                blank_count += 1
                if blank_count == 2:
                    if detail_rows.get(current_name):
                        detail_rows[current_name].append({"group_boundary": True})
                elif blank_count == 3:
                    pending_blank = False
                    if detail_rows.get(current_name):
                        state = "SEEKING_TOTAL"
                    else:
                        blank_count = 0
                continue

            blank_count = 0

            # Column header row appearing at the start of a new section
            new_map = detect_column_map(row)
            if new_map:
                column_map = new_map
                debug(f"Re-detected column map for '{current_name}': {column_map}")
                continue

            label = None
            from_col_a = False
            col_a_val = row[0]
            if col_a_val is not None and str(col_a_val).strip():
                label = str(col_a_val).strip()
                from_col_a = True
            else:
                for c in row[1:4]:
                    if c is not None and str(c).strip():
                        label = str(c).strip()
                        break

            if not label:
                continue

            if from_col_a and re.match(r'^RE\s*:', label, re.IGNORECASE):
                debug(f"  [skip RE] '{label}'")
                continue

            values = {}
            for bucket in BUCKETS:
                raw = row[column_map[bucket]] if bucket in column_map else None
                values[bucket] = float(raw) if raw is not None else 0.0

            if from_col_a and label and label[0].isdigit():
                label = strip_m_prefix(label)

            row_type = "matter" if (from_col_a and label[0].isdigit()) else "subtotal"
            if row_type == "subtotal" and not label.lower().endswith("total"):
                label = label + " Total"
            debug(f"  [{row_type}] '{label}'")
            detail_rows[current_name].append({"label": label, "values": values, "type": row_type})

        elif state == "SEEKING_TOTAL":
            if is_blank_row(row):
                continue

            label = None
            col_a_val = row[0]
            if col_a_val is not None and str(col_a_val).strip():
                label = str(col_a_val).strip()
            else:
                for c in row[1:4]:
                    if c is not None and str(c).strip():
                        label = str(c).strip()
                        break
            if not label:
                label = f"{current_name} TOTAL"

            values = {}
            for bucket in BUCKETS:
                raw = row[column_map[bucket]] if bucket in column_map else None
                values[bucket] = float(raw) if raw is not None else 0.0

            debug(f"  [total] '{label}'")
            detail_rows[current_name].append({"label": label, "values": values, "type": "total"})

            # Advance to next timekeeper
            timekeeper_index += 1
            current_name = get_timekeeper_name(timekeeper_index)
            attorneys.append(current_name)
            detail_rows[current_name] = []
            blank_count = 0
            debug(f"Starting timekeeper {timekeeper_index}: '{current_name}'")
            state = "READING_MATTERS"

    attorneys = [a for a in attorneys if detail_rows.get(a)]
    for name in list(detail_rows.keys()):
        if not detail_rows[name]:
            del detail_rows[name]

    return attorneys, detail_rows, header_info


def resolve_blanks(entries):

    groups = []
    current = []
    for entry in entries:
        if entry.get("group_boundary"):
            if current:
                groups.append(current)
            current = []
        else:
            current.append(entry)
    if current:
        groups.append(current)

    def group_bal_due(group):
        subtotals = [e for e in group if e.get("type") == "subtotal"]
        if subtotals:
            return subtotals[-1]["values"].get("Bal Due", 0.0)
        return sum(e["values"].get("Bal Due", 0.0) for e in group if e.get("type") == "matter")

    # Drop groups whose total Bal Due is negative
    groups = [g for g in groups if group_bal_due(g) >= 0]

    def is_multi(group):
        return any(e.get("type") == "subtotal" for e in group)

    result = []
    for i, group in enumerate(groups):
        if i > 0 and (is_multi(groups[i - 1]) or is_multi(group)):
            result.append({"blank": True})
        result.extend(group)
    return result


def write_detail_sheet(ws, attorneys, detail_rows, header_info):

    ACCT_FMT = r'_($* #,##0.00_);_($* \(#,##0.00\);_($* \-??_);_(@_)'
    DATE_FMT = "m/d/yyyy"
    grey_fill = PatternFill(fill_type="solid", fgColor="A5A5A5")
    center = Alignment(horizontal="center")
    right  = Alignment(horizontal="right")

    BAL_DUE_COL = BUCKET_START_COL + len(BUCKETS) - 1          # col L
    AGING_END_COL = get_column_letter(BUCKET_START_COL + len(BUCKETS) - 2)  # K
    AGING_START_COL = get_column_letter(BUCKET_START_COL)       # F


    col_widths = {"A": 56.29}
    for i in range(len(BUCKETS)):
        col_widths[get_column_letter(BUCKET_START_COL + i)] = [
            16.43, 16.0, 16.14, 17.86, 17.15, 19.0, 21.71
        ][i]
    col_widths[get_column_letter(BUCKET_START_COL + len(BUCKETS))]     = 49.71  # ATTORNEY NOTES
    col_widths[get_column_letter(BUCKET_START_COL + len(BUCKETS) + 1)] = 35.0   # ACCOUNTING NOTES
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    def set_row_height(row, height):
        ws.row_dimensions[row].height = height

    def data_cell(row, col, value, font, number_format=None, alignment=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        if number_format:
            c.number_format = number_format
        if alignment:
            c.alignment = alignment
        return c

    current_row = 1

    set_row_height(1, 22.5)
    date_val = header_info.get("date")
    if date_val:
        c = data_cell(1, 1, date_val, Font(name="Arial", size=10))
        c.number_format = DATE_FMT
    title_val = header_info.get("title", "Tabs3 Summary Accounts Receivable Report")
    data_cell(1, BUCKET_START_COL, title_val,
              Font(name="Arial", size=15, bold=True), alignment=center)
    current_row = 2

    set_row_height(2, 24.0)
    name_label = header_info.get("name_label")
    if name_label:
        data_cell(2, BUCKET_START_COL, name_label,
                  Font(name="Arial", size=14, bold=True), alignment=center)
    current_row = 3

    set_row_height(3, 12.75)
    current_row = 4

    first_section = True

    for name in attorneys:
        entries = resolve_blanks(detail_rows.get(name, []))

        if not first_section:
            set_row_height(current_row, 24.75)
            current_row += 1
        first_section = False

        set_row_height(current_row, 24.75)
        data_cell(current_row, 1, name, Font(name="Arial", size=12, bold=True))

        for i, bucket in enumerate(BUCKETS[:-1]):   # 0-30 … 151+
            c = ws.cell(row=current_row, column=BUCKET_START_COL + i, value=bucket)
            c.font   = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            c.fill   = grey_fill
            c.alignment = center
        # Bal Due header
        c = ws.cell(row=current_row, column=BAL_DUE_COL, value="Bal Due")
        c.font   = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c.fill   = grey_fill
        c.alignment = center
        # Notes headers
        notes_labels = ["ATTORNEY NOTES", "ACCOUNTING NOTES"]
        for i, label in enumerate(notes_labels):
            c = ws.cell(row=current_row, column=BAL_DUE_COL + 1 + i, value=label)
            c.font   = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
            c.fill   = grey_fill
            c.alignment = center
        current_row += 1

        # ---- Data rows ----
        for entry in entries:
            if entry.get("blank"):
                set_row_height(current_row, 24.75)
                current_row += 1
                continue

            set_row_height(current_row, 24.75)
            row_type = entry.get("type", "matter")
            lbl      = entry["label"]
            vals     = entry["values"]

            if row_type == "matter":
                label_font = Font(name="Arial", size=10)
                value_font = Font(name="Arial", size=11)
                lbl_align  = None
            elif row_type == "subtotal":
                label_font = Font(name="Arial", size=11)
                value_font = Font(name="Arial", size=11)
                lbl_align  = right
            else:  # total
                label_font = Font(name="Arial", size=12, bold=True)
                value_font = Font(name="Arial", size=12, bold=True)
                lbl_align  = right

            data_cell(current_row, 1, lbl, label_font, alignment=lbl_align)

            # Aging bucket columns (0-30 … 151+) — raw values
            for i, bucket in enumerate(BUCKETS[:-1]):
                data_cell(current_row, BUCKET_START_COL + i,
                          vals[bucket], value_font, ACCT_FMT)

            # Bal Due — formula that sums the 6 aging columns on this row
            formula = f"=SUM({AGING_START_COL}{current_row}:{AGING_END_COL}{current_row})"
            data_cell(current_row, BAL_DUE_COL, formula, value_font, ACCT_FMT)

            current_row += 1


AR_SHEET_NAME     = "Tabs3 Summary Accounts Receivable"
RECAP_SHEET_NAME  = "Recap Sheet"


def find_ar_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("Tabs3 Summary"):
            return wb[name]
    return None


def find_recap_sheet(wb):
    for name in wb.sheetnames:
        if name.lower().startswith("recap"):
            return wb[name]
    return None


def get_prev_matter_ids(wb):

    ws = find_ar_sheet(wb)
    if not ws:
        return set()
    ids = set()
    for row in ws.iter_rows(values_only=True):
        label = row[0]
        if label and str(label).strip():
            m = re.match(r'^(\d[\d.]*)', str(label).strip())
            if m:
                ids.add(m.group(1))
    return ids


def replace_sheet(wb, old_sheet_name, new_name):

    if old_sheet_name in wb.sheetnames:
        pos = wb.sheetnames.index(old_sheet_name)
        del wb[old_sheet_name]
    else:
        pos = 0
    return wb.create_sheet(new_name, pos)


def extract_fill(cell):

    HEADER_GREY = "FFA5A5A5"
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if not color:
        return None
    try:
        if color.type == "rgb":
            rgb = color.rgb
            if not rgb or rgb in ("00000000", HEADER_GREY):
                return None
            return PatternFill(fill_type="solid", fgColor=rgb)
        elif color.type == "theme":
            from openpyxl.styles.colors import Color
            new_color = Color(theme=int(color.theme), tint=float(color.tint or 0))
            return PatternFill(fill_type="solid", fgColor=new_color)
        elif color.type == "indexed":
            from openpyxl.styles.colors import Color
            new_color = Color(indexed=int(color.indexed))
            return PatternFill(fill_type="solid", fgColor=new_color)
    except Exception:
        pass
    return None


def parse_previous_comments(wb):

    ws = find_ar_sheet(wb)
    if ws is None:
        print("Could not find AR sheet in previous file.", file=sys.stderr)
        return {}

    atty_col = acct_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "ATTORNEY NOTES":
                atty_col = cell.column
            elif cell.value == "ACCOUNTING NOTES":
                acct_col = cell.column
        if atty_col or acct_col:
            break

    if not atty_col and not acct_col:
        print("Could not find ATTORNEY NOTES / ACCOUNTING NOTES columns in previous file.",
              file=sys.stderr)
        return {}

    comments = {}
    for row in ws.iter_rows():
        label_val = row[0].value
        if not label_val:
            continue
        m = re.match(r'^(\d[\d.]*)', str(label_val).strip())
        if not m:
            continue
        matter_id = m.group(1)

        entry = {"atty_note": None, "atty_fill": None, "acct_note": None, "acct_fill": None}
        if atty_col:
            c = ws.cell(row=row[0].row, column=atty_col)
            if c.value and str(c.value).strip():
                entry["atty_note"] = str(c.value).strip()
                entry["atty_fill"] = extract_fill(c)
        if acct_col:
            c = ws.cell(row=row[0].row, column=acct_col)
            if c.value and str(c.value).strip():
                entry["acct_note"] = str(c.value).strip()
                entry["acct_fill"] = extract_fill(c)

        if entry["atty_note"] or entry["acct_note"]:
            comments[matter_id] = entry

    return comments


def apply_comments(ws, comments):
    if not comments:
        return

    atty_col = acct_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "ATTORNEY NOTES":
                atty_col = cell.column
            elif cell.value == "ACCOUNTING NOTES":
                acct_col = cell.column
        if atty_col or acct_col:
            break

    if not atty_col and not acct_col:
        return

    last_col = acct_col or atty_col

    carried = 0
    for row in ws.iter_rows():
        label_val = row[0].value
        if not label_val:
            continue
        m = re.match(r'^(\d[\d.]*)', str(label_val).strip())
        if not m:
            continue
        matter_id = m.group(1)
        if matter_id not in comments:
            continue

        entry = comments[matter_id]
        row_num = row[0].row

        # Write note text
        if atty_col and entry["atty_note"]:
            ws.cell(row=row_num, column=atty_col, value=entry["atty_note"])
        if acct_col and entry["acct_note"]:
            ws.cell(row=row_num, column=acct_col, value=entry["acct_note"])

        # Determine row highlight color: attorney fill takes priority over accounting fill
        row_fill = entry["atty_fill"] or entry["acct_fill"]
        if row_fill:
            for col in range(1, last_col + 1):
                ws.cell(row=row_num, column=col).fill = row_fill

        carried += 1

    print(f"Carried over comments for {carried} matter(s) from previous file.")


def write_recap_sheet(ws, attorneys, detail_rows, header_info):
    ACCT_FMT = r'_($* #,##0.00_);_($* \(#,##0.00\);_($* \-??_);_(@_)'
    grey_fill = PatternFill(fill_type="solid", fgColor="A5A5A5")
    center    = Alignment(horizontal="center")
    right     = Alignment(horizontal="right")

    BAL_DUE_COL    = BUCKET_START_COL + len(BUCKETS) - 1
    AGING_END_COL  = get_column_letter(BUCKET_START_COL + len(BUCKETS) - 2)
    AGING_START_COL = get_column_letter(BUCKET_START_COL)

    ws.column_dimensions["A"].width = 20
    for i in range(len(BUCKETS)):
        ws.column_dimensions[get_column_letter(BUCKET_START_COL + i)].width = [
            16.43, 16.0, 16.14, 17.86, 17.15, 19.0, 21.71][i]

    ws.row_dimensions[1].height = 22.5
    c = ws.cell(row=1, column=BUCKET_START_COL,
                value=header_info.get("title", "Tabs3 Summary Accounts Receivable Report"))
    c.font = Font(name="Arial", size=15, bold=True)
    c.alignment = center

    ws.row_dimensions[2].height = 24.0
    if header_info.get("name_label"):
        c = ws.cell(row=2, column=BUCKET_START_COL, value=header_info["name_label"])
        c.font = Font(name="Arial", size=14, bold=True)
        c.alignment = center

    ws.row_dimensions[3].height = 12.75

    ws.row_dimensions[4].height = 24.75
    for i, bucket in enumerate(BUCKETS[:-1]):
        c = ws.cell(row=4, column=BUCKET_START_COL + i, value=bucket)
        c.font  = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c.fill  = grey_fill
        c.alignment = center
    c = ws.cell(row=4, column=BAL_DUE_COL, value="Bal Due")
    c.font  = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    c.fill  = grey_fill
    c.alignment = center

    current_row = 5

    for name in attorneys:
        entries = detail_rows.get(name, [])
        total_entry = next((e for e in entries if e.get("type") == "total"), None)
        if total_entry:
            values = total_entry["values"]
        else:

            values = {b: 0.0 for b in BUCKETS}
            for e in entries:
                if isinstance(e, dict) and e.get("type") in ("matter", "subtotal"):
                    for b in BUCKETS:
                        values[b] += e["values"].get(b, 0.0)

        ws.row_dimensions[current_row].height = 24.75
        ws.cell(row=current_row, column=1, value=name).font = Font(name="Arial", size=12, bold=True)

        for i, bucket in enumerate(BUCKETS[:-1]):
            c = ws.cell(row=current_row, column=BUCKET_START_COL + i, value=values[bucket])
            c.font = Font(name="Arial", size=11)
            c.number_format = ACCT_FMT

        c = ws.cell(row=current_row, column=BAL_DUE_COL,
                    value=f"=SUM({AGING_START_COL}{current_row}:{AGING_END_COL}{current_row})")
        c.font = Font(name="Arial", size=11)
        c.number_format = ACCT_FMT
        current_row += 1

    data_start = 5
    data_end   = current_row - 1
    ws.row_dimensions[current_row].height = 24.75
    c = ws.cell(row=current_row, column=1, value="Total")
    c.font = Font(name="Arial", size=12, bold=True)
    c.alignment = right

    for i in range(len(BUCKETS) - 1):
        col_letter = get_column_letter(BUCKET_START_COL + i)
        c = ws.cell(row=current_row, column=BUCKET_START_COL + i,
                    value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        c.font = Font(name="Arial", size=11, bold=True)
        c.number_format = ACCT_FMT

    c = ws.cell(row=current_row, column=BAL_DUE_COL,
                value=f"=SUM({AGING_START_COL}{current_row}:{AGING_END_COL}{current_row})")
    c.font = Font(name="Arial", size=11, bold=True)
    c.number_format = ACCT_FMT


def main():
    global DEBUG

    args = sys.argv[1:]
    if "--debug" in args:
        DEBUG = True
        args = [a for a in args if a != "--debug"]

    prev_path = None
    if "--prev" in args:
        idx = args.index("--prev")
        if idx + 1 >= len(args):
            print("Error: --prev requires a file path argument.", file=sys.stderr)
            sys.exit(1)
        prev_path = args[idx + 1]
        args = args[:idx] + args[idx + 2:]

    if len(args) != 2:
        print("Usage: python transform_ar.py <input.xlsx> <output.xlsx> [--prev <prev_output.xlsx>] [--debug]")
        sys.exit(1)

    input_path, output_path = args[0], args[1]

    try:
        wb_in = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    except FileNotFoundError:
        print(f"Error: Input file not found: '{input_path}'", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: Could not open input file: {e}", file=sys.stderr)
        sys.exit(1)

    ws_in = wb_in.active
    attorneys, detail_rows, header_info = parse_input(ws_in)

    prev_comments = {}

    if prev_path:
        try:
            # Load the previous workbook — this becomes the base for output
            wb_out = openpyxl.load_workbook(prev_path)
        except FileNotFoundError:
            print(f"Error: Previous file not found: '{prev_path}'", file=sys.stderr)
            sys.exit(1)
        except Exception as e:
            print(f"Error: Could not open previous file: {e}", file=sys.stderr)
            sys.exit(1)

        # Extract comments before modifying any sheets
        prev_comments = parse_previous_comments(wb_out)
        print(f"Loaded {len(prev_comments)} comment(s) from previous file.")

        ar_old   = find_ar_sheet(wb_out)
        ar_name  = ar_old.title if ar_old else AR_SHEET_NAME
        rec_old  = find_recap_sheet(wb_out)
        rec_name = rec_old.title if rec_old else RECAP_SHEET_NAME

        ws_detail = replace_sheet(wb_out, ar_name,  ar_name)
        ws_recap  = replace_sheet(wb_out, rec_name, rec_name)
    else:
        wb_out    = Workbook()
        ws_detail = wb_out.active
        ws_detail.title = AR_SHEET_NAME
        ws_recap  = wb_out.create_sheet(RECAP_SHEET_NAME)

    write_detail_sheet(ws_detail, attorneys, detail_rows, header_info)
    write_recap_sheet(ws_recap, attorneys, detail_rows, header_info)

    if prev_comments:
        apply_comments(ws_detail, prev_comments)

    try:
        wb_out.save(output_path)
        print(f"Output saved to: {output_path}")
        print(f"Processed {len(attorneys)} timekeeper(s): {', '.join(attorneys)}")
    except Exception as e:
        print(f"Error: Could not save output file: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
